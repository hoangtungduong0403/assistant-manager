"""
Phân tích công nợ - Mission: Đánh giá công nợ (Debt Evaluation)
================================================================
USAGE:
    python3 analyze_debt.py <SGA_Revenue_and_Debt.xlsx> <output_dir>

Đọc sheet DEBT (kỳ báo cáo lấy trực tiếp từ ô E7/G7 - đổi kỳ bằng cách sửa
2 ô này trên Google Sheet gốc rồi tải lại file, KHÔNG cần sửa script).

Vì sheet gốc có lỗi công thức đã biết (vùng SUM ở dòng Tổng cộng không nhất
quán - xem mapping.md), script này KHÔNG tin vào các ô Tổng cộng có sẵn mà
luôn cộng tay trực tiếp từ dữ liệu chi tiết (dòng 14 trở đi). Nếu số tự tính
lệch với số hiển thị trên sheet, lệch đó được ghi lại trong analysis.json
(report_total_g / true_total_g / diff_g) để đưa vào báo cáo mục 2.

OUTPUT:
    <output_dir>/analysis.json           - toàn bộ số liệu đã tính cho build_report.js
    <output_dir>/charts/reconciliation.png
    <output_dir>/charts/top_debtors.png
    <output_dir>/charts/top_credit.png
    <output_dir>/charts/concentration.png
"""

import sys
import json
import os
from datetime import datetime

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

plt.rcParams["font.family"] = "DejaVu Sans"

SHEET_NAME = "DEBT"
DATA_START_ROW = 14
TOP_N = 10
CONCENTRATION_THRESHOLD = 20.0  # % - ngưỡng cảnh báo tập trung vào 1 khách hàng


def fmt_vn_date(d):
    if d is None:
        return None
    return d.strftime("%d/%m/%Y")


def load_rows(ws):
    rows = []
    r = DATA_START_ROW
    while True:
        ma = ws.cell(row=r, column=2).value
        if ma in (None, ""):
            # allow a few blank rows in case of gaps, but stop after 3 consecutive blanks
            blanks = 0
            rr = r
            while blanks < 3:
                v = ws.cell(row=rr, column=2).value
                if v in (None, ""):
                    blanks += 1
                    rr += 1
                else:
                    break
            if blanks >= 3:
                break
            r += 1
            continue
        rows.append({
            "stt": ws.cell(row=r, column=1).value,
            "ma": str(ma).strip(),
            "mst": ws.cell(row=r, column=3).value,
            "dau_ky_thu": ws.cell(row=r, column=5).value or 0,
            "dau_ky_tra": ws.cell(row=r, column=6).value or 0,
            "trong_ky_thu": ws.cell(row=r, column=7).value or 0,
            "trong_ky_tra": ws.cell(row=r, column=8).value or 0,
            "cuoi_ky_thu": ws.cell(row=r, column=9).value or 0,
            "cuoi_ky_tra": ws.cell(row=r, column=10).value or 0,
            "ghi_chu": ws.cell(row=r, column=11).value,
            "kho_doi": ws.cell(row=r, column=12).value,
            "han": ws.cell(row=r, column=13).value,
        })
        r += 1
    return rows


def hbar_chart(path, labels, values, color):
    labels = labels[::-1]
    values = values[::-1]
    fig, ax = plt.subplots(figsize=(7.2, 3.6), dpi=200)
    bars = ax.barh(labels, [v / 1e6 for v in values], color=color, height=0.6)
    ax.set_xlabel("Triệu đồng", fontsize=10, color="#5F5E5A")
    ax.tick_params(axis="y", labelsize=10)
    ax.tick_params(axis="x", labelsize=9, colors="#5F5E5A")
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#c3c2b7")
    ax.spines["bottom"].set_color("#c3c2b7")
    ax.grid(axis="x", color="#e1e0d9", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for bar, v in zip(bars, values):
        ax.text(bar.get_width() + 1, bar.get_y() + bar.get_height() / 2, f"{v/1e6:.1f}tr", va="center", fontsize=9, color="#2C2C2A")
    fig.tight_layout()
    fig.savefig(path, transparent=True)
    plt.close(fig)


def concentration_chart(path, top_pct, rest_pct, n_top, n_rest):
    fig, ax = plt.subplots(figsize=(4.2, 4.2), dpi=200)
    colors = ["#d03b3b", "#e1e0d9"]
    wedges, texts, autotexts = ax.pie(
        [top_pct, rest_pct], colors=colors, startangle=90, counterclock=False,
        wedgeprops=dict(width=0.42, edgecolor="white", linewidth=2),
        autopct=lambda p: f"{p:.1f}%", pctdistance=0.79,
    )
    for t in autotexts:
        t.set_fontsize(11)
        t.set_color("#2C2C2A")
    ax.legend(wedges, [f"Top {n_top} khách hàng", f"{n_rest} khách hàng còn lại"],
              loc="lower center", bbox_to_anchor=(0.5, -0.18), fontsize=9, frameon=False, ncol=1)
    ax.set_title("Mức tập trung công nợ phải thu", fontsize=12, color="#2C2C2A", pad=10)
    fig.tight_layout()
    fig.savefig(path, transparent=True)
    plt.close(fig)


def reconciliation_chart(path, dau_ky, tang, thu, cuoi_ky):
    fig, ax = plt.subplots(figsize=(7.2, 3.2), dpi=200)
    cats = ["Đầu kỳ", "+ Phát sinh tăng", "- Đã thu", "Cuối kỳ"]
    vals = [dau_ky / 1e6, tang / 1e6, -thu / 1e6, cuoi_ky / 1e6]
    bar_colors = ["#888780", "#639922", "#d03b3b", "#888780"]
    bars = ax.bar(cats, vals, color=bar_colors, width=0.55, zorder=3)
    ax.axhline(0, color="#c3c2b7", linewidth=0.8)
    ax.set_ylabel("Triệu đồng", fontsize=10, color="#5F5E5A")
    ax.tick_params(axis="x", labelsize=10)
    ax.tick_params(axis="y", labelsize=9, colors="#5F5E5A")
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.grid(axis="y", color="#e1e0d9", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for bar, v in zip(bars, vals):
        va = "bottom" if v >= 0 else "top"
        off = 8 if v >= 0 else -8
        ax.annotate(f"{v:.1f}tr", (bar.get_x() + bar.get_width() / 2, v), textcoords="offset points", xytext=(0, off), ha="center", fontsize=9, color="#2C2C2A")
    fig.tight_layout()
    fig.savefig(path, transparent=True)
    plt.close(fig)


def main(xlsx_path, out_dir):
    charts_dir = os.path.join(out_dir, "charts")
    os.makedirs(charts_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]

    ky_tu = ws["E7"].value
    ky_den = ws["G7"].value

    rows = load_rows(ws)
    n_total = len(rows)

    true_dau_ky_thu = sum(r["dau_ky_thu"] for r in rows)
    true_dau_ky_tra = sum(r["dau_ky_tra"] for r in rows)
    true_trong_ky_thu = sum(r["trong_ky_thu"] for r in rows)
    true_trong_ky_tra = sum(r["trong_ky_tra"] for r in rows)
    true_cuoi_ky_thu = sum(r["cuoi_ky_thu"] for r in rows)
    true_cuoi_ky_tra = sum(r["cuoi_ky_tra"] for r in rows)

    # So sánh với ô Tổng cộng có sẵn trên sheet (dòng 12) để phát hiện lệch do lỗi công thức
    report_totals = {
        "dau_ky_thu": ws["E12"].value or 0,
        "dau_ky_tra": ws["F12"].value or 0,
        "trong_ky_thu": ws["G12"].value or 0,
        "trong_ky_tra": ws["H12"].value or 0,
        "cuoi_ky_thu": ws["I12"].value or 0,
        "cuoi_ky_tra": ws["J12"].value or 0,
    }
    true_totals = {
        "dau_ky_thu": true_dau_ky_thu, "dau_ky_tra": true_dau_ky_tra,
        "trong_ky_thu": true_trong_ky_thu, "trong_ky_tra": true_trong_ky_tra,
        "cuoi_ky_thu": true_cuoi_ky_thu, "cuoi_ky_tra": true_cuoi_ky_tra,
    }
    discrepancies = []
    for key in true_totals:
        diff = round(true_totals[key] - report_totals[key], 2)
        if abs(diff) >= 1:  # bỏ qua sai số làm tròn < 1đ
            discrepancies.append({
                "field": key, "report_value": report_totals[key],
                "true_value": true_totals[key], "diff": diff,
                "diff_pct": round(diff / true_totals[key] * 100, 1) if true_totals[key] else None,
            })

    # Top N khách nợ nhiều nhất (cuối kỳ - phải thu)
    debtors = sorted([r for r in rows if r["cuoi_ky_thu"] > 0], key=lambda r: -r["cuoi_ky_thu"])
    top_debtors = debtors[:TOP_N]
    top_debtors_sum = sum(r["cuoi_ky_thu"] for r in top_debtors)
    top_debtors_pct = round(top_debtors_sum / true_cuoi_ky_thu * 100, 1) if true_cuoi_ky_thu else 0

    # Top N khách trả trước / dư có nhiều nhất
    creditors = sorted([r for r in rows if r["cuoi_ky_tra"] > 0], key=lambda r: -r["cuoi_ky_tra"])
    top_creditors = creditors[:TOP_N]
    top_creditors_sum = sum(r["cuoi_ky_tra"] for r in top_creditors)
    top_creditors_pct = round(top_creditors_sum / true_cuoi_ky_tra * 100, 1) if true_cuoi_ky_tra else 0

    # Khách hàng vượt ngưỡng tập trung cá nhân (>20% tổng công nợ)
    concentrated = [r for r in debtors if true_cuoi_ky_thu and (r["cuoi_ky_thu"] / true_cuoi_ky_thu * 100) >= CONCENTRATION_THRESHOLD]

    # Nợ khó đòi
    kho_doi_flagged = [r for r in rows if (r["kho_doi"] or "").strip() in ("Nợ khó đòi", "Khó đòi")]
    kho_doi_active = [r for r in kho_doi_flagged if r["cuoi_ky_thu"] > 0]
    kho_doi_stale = [r for r in kho_doi_flagged if r["cuoi_ky_thu"] == 0]
    kho_doi_active_sum = sum(r["cuoi_ky_thu"] for r in kho_doi_active)

    # Ghi chú "Sai số" hoặc các ghi chú cần rà soát thủ công
    flagged_notes = [r for r in rows if (r["ghi_chu"] or "").strip() not in ("", None)]

    # Dữ liệu "Thời hạn thanh toán" có đủ để làm aging không
    han_filled = sum(1 for r in rows if r["han"] is not None)
    han_coverage_pct = round(han_filled / n_total * 100, 1) if n_total else 0

    result = {
        "ky_tu": fmt_vn_date(ky_tu),
        "ky_den": fmt_vn_date(ky_den),
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "n_companies_total": n_total,
        "n_debtors": len(debtors),
        "n_creditors": len(creditors),
        "totals": true_totals,
        "report_totals": report_totals,
        "discrepancies": discrepancies,
        "top_debtors": [{"ma": r["ma"], "so_tien": r["cuoi_ky_thu"]} for r in top_debtors],
        "top_debtors_sum": top_debtors_sum,
        "top_debtors_pct": top_debtors_pct,
        "top_creditors": [{"ma": r["ma"], "so_tien": r["cuoi_ky_tra"]} for r in top_creditors],
        "top_creditors_sum": top_creditors_sum,
        "top_creditors_pct": top_creditors_pct,
        "concentration_threshold_pct": CONCENTRATION_THRESHOLD,
        "concentrated_customers": [{"ma": r["ma"], "pct": round(r["cuoi_ky_thu"] / true_cuoi_ky_thu * 100, 1)} for r in concentrated],
        "kho_doi_flagged_count": len(kho_doi_flagged),
        "kho_doi_active": [{"ma": r["ma"], "so_tien": r["cuoi_ky_thu"]} for r in kho_doi_active],
        "kho_doi_active_sum": kho_doi_active_sum,
        "kho_doi_stale_count": len(kho_doi_stale),
        "flagged_notes": [{"ma": r["ma"], "ghi_chu": r["ghi_chu"]} for r in flagged_notes],
        "han_coverage_pct": han_coverage_pct,
    }

    with open(os.path.join(out_dir, "analysis.json"), "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    # Charts
    reconciliation_chart(
        os.path.join(charts_dir, "reconciliation.png"),
        true_dau_ky_thu, true_trong_ky_thu, true_trong_ky_tra, true_cuoi_ky_thu,
    )
    hbar_chart(os.path.join(charts_dir, "top_debtors.png"),
               [r["ma"] for r in top_debtors], [r["cuoi_ky_thu"] for r in top_debtors], "#d03b3b")
    hbar_chart(os.path.join(charts_dir, "top_credit.png"),
               [r["ma"] for r in top_creditors], [r["cuoi_ky_tra"] for r in top_creditors], "#1baf7a")
    concentration_chart(os.path.join(charts_dir, "concentration.png"),
                         top_debtors_pct, round(100 - top_debtors_pct, 1), TOP_N, max(len(debtors) - TOP_N, 0))

    print(f"Kỳ báo cáo: {result['ky_tu']} - {result['ky_den']}")
    print(f"Số công ty có phát sinh: {n_total}")
    print(f"Phát hiện {len(discrepancies)} chênh lệch giữa số hiển thị trên sheet và số tính đúng.")
    print(f"Đã ghi analysis.json và 4 biểu đồ vào {out_dir}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 analyze_debt.py <SGA_Revenue_and_Debt.xlsx> <output_dir>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
