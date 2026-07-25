"""
Update Summary Quotation from Client Information
===================================================
Mission: Bao gia (Quotation) - SGA

USAGE
-----
    python3 update_summary_quotation.py <client_information.xlsx> <summary_quotation_template.xlsx>

Reads every customer row from the Client Information file and adds one new
row per customer into the correct sheet(s) of the Summary Quotation
(DATA LONG - TERM and/or DATA SHORT - TERM). Saves the result as a NEW file
named SummaryQuotation_<yyyyMMdd_HHmm>.xlsx -- the template itself is never
modified.

CLASSIFICATION LOGIC (updated 2026-07-25)
------------------------------------------
Long-term / Short-term is now decided from the actual VALUES in columns
O / P / Q, not from the LOAI column (N). LOAI is only used as a fallback
when O, P and Q are all blank. See mapping.md for the full rule table.

    - O (Dich vu phan tich - thong ke) has a value      -> customer uses
      LONG-TERM. The value itself (Standard/Premium/Platinum) is just the
      package level, written into DATA LONG - TERM column D.
    - P (Dich vu nhan su) or Q (Dich vu phap che doanh nghiep) has a value
      -> customer uses SHORT-TERM. Package levels from P/Q are joined into
      the DIEN GIAI text of DATA SHORT - TERM column G.
    - O has a value AND (P or Q) has a value -> customer uses BOTH. Two
      separate rows are added, one per sheet. The SHORT-TERM row's DIEN
      GIAI only includes P/Q services, never O.
    - O, P, Q all blank, but LOAI resolves to "Long-term" or "Short-term"
      -> fallback: add one row to the corresponding sheet, with the
      package/DIEN GIAI field left blank ("Chua duoc cung cap").
    - O, P, Q all blank AND LOAI also blank/unrecognized -> customer is
      skipped entirely and listed separately for manual review. Never
      guessed.

Duplicate-checking (by MA CONG TY) is done INDEPENDENTLY per sheet: a
customer already present in DATA LONG - TERM but not yet in
DATA SHORT - TERM will still get a new row added to DATA SHORT - TERM (and
vice versa) if the classification calls for it.

OTHER ASSUMPTIONS / DEFAULTS (documented so they can be reviewed each run)
----------------------------------------------------------------------
- MA SO THUE (tax code) is written as a static value (from Client Info),
  not the original IMPORTRANGE formula, since IMPORTRANGE does not work in
  a local .xlsx file.
- "So hop dong" (DATA LONG - TERM, col E) defaults to "CHUA CO" when no
  contract exists yet, per mapping.md.
- Sequential numbers (col G in LONG-TERM, col D in SHORT-TERM) auto-continue
  from the highest existing number in that sheet.
- VAT (SHORT-TERM col I) defaults to 0.08 (8%), the standing default noted
  in mapping.md.
- Fields with NO source in Client Information are left BLANK, never guessed:
    LONG-TERM:  F (Tinh trang hop dong), I (DVT phu phi), J (Hoa don),
                L (Ngay thuc hien), N (Tinh trang bao gia)
    SHORT-TERM: F (Ngay thuc hien), H (Don gia), M (Tinh trang)
  Because "Ngay hieu luc" is an existing formula that adds 15 days to a
  blank "Ngay thuc hien", it will show a meaningless placeholder date
  (15/01/1900) until that field is filled in manually -- this is a known
  side effect of the ORIGINAL formula, left untouched per convention.
- Style (font/fill/border/number format) is copied from the last existing
  data row in each sheet, so new rows visually match the template.

Run `python3 /mnt/skills/public/xlsx/scripts/recalc.py <output.xlsx>` after
this script to compute formula values before delivering the file.
"""

import os
import sys
import re
from copy import copy
from datetime import datetime

import openpyxl

CLIENT_INFO_HEADER_ROW = 3
CLIENT_INFO_DATA_START_ROW = 4
COL_MA_CONG_TY = 2       # B
COL_MA_SO_THUE = 5       # E
COL_LOAI = 14             # N
COL_SERVICE_O = 15        # O - Dich vu phan tich - thong ke (Long-term)
COL_SERVICE_P = 16        # P - Dich vu nhan su (Short-term)
COL_SERVICE_Q = 17        # Q - Dich vu phap che doanh nghiep (Short-term)
SERVICE_NAMES = {
    COL_SERVICE_O: "Dịch vụ phân tích - thống kê",
    COL_SERVICE_P: "Dịch vụ nhân sự",
    COL_SERVICE_Q: "Dịch vụ pháp chế doanh nghiệp",
}

LT_SHEET = "DATA LONG - TERM"
ST_SHEET = "DATA SHORT - TERM"


def copy_row_style(ws, src_row, dst_row, ncols):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, ncols + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        d.font = copy(s.font)
        d.fill = copy(s.fill)
        d.border = copy(s.border)
        d.number_format = s.number_format
        d.alignment = copy(s.alignment)
        d.protection = copy(s.protection)


def read_client_information(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Information"]
    customers = []
    r = CLIENT_INFO_DATA_START_ROW
    while True:
        code = ws.cell(row=r, column=COL_MA_CONG_TY).value
        if code in (None, ""):
            break
        mst = ws.cell(row=r, column=COL_MA_SO_THUE).value
        loai = ws.cell(row=r, column=COL_LOAI).value

        def cell_val(col):
            v = ws.cell(row=r, column=col).value
            return str(v).strip() if v not in (None, "") else None

        o_level = cell_val(COL_SERVICE_O)
        p_level = cell_val(COL_SERVICE_P)
        q_level = cell_val(COL_SERVICE_Q)

        customers.append({
            "code": str(code).strip(),
            "mst": str(mst).strip() if mst not in (None, "") else "",
            "loai": str(loai).strip() if loai not in (None, "") else "",
            "o_level": o_level,
            "p_level": p_level,
            "q_level": q_level,
        })
        r += 1
    return customers


def classify(cust):
    """Return (uses_long, uses_short, via_fallback) for a customer, per the
    O/P/Q-based rules documented at the top of this file."""
    has_o = cust["o_level"] is not None
    has_short_service = cust["p_level"] is not None or cust["q_level"] is not None

    if has_o or has_short_service:
        return has_o, has_short_service, False

    loai_lower = cust["loai"].lower()
    if "long-term" in loai_lower:
        return True, False, True
    if "short-term" in loai_lower:
        return False, True, True

    return False, False, False  # unclassifiable


def existing_codes(ws, code_col, start_row=4):
    codes = set()
    last_row = start_row - 1
    r = start_row
    while r <= ws.max_row:
        v = ws.cell(row=r, column=code_col).value
        if v not in (None, ""):
            codes.add(str(v).strip().upper())
            last_row = r
        r += 1
    return codes, last_row


def next_seq(ws, col, start_row, last_row, width=3):
    max_n = 0
    for r in range(start_row, last_row + 1):
        v = ws.cell(row=r, column=col).value
        if v not in (None, ""):
            m = re.match(r"^\d+", str(v))
            if m:
                max_n = max(max_n, int(m.group()))
    return str(max_n + 1).zfill(width)


def add_long_term_row(ws, r, cust, via_fallback):
    level = cust["o_level"] if cust["o_level"] else ""
    seq = next_seq(ws, 7, 4, r - 1)
    ws.cell(row=r, column=1, value=f'=IF(B{r}="","",SUBTOTAL(3,$B$4:B{r}))')
    ws.cell(row=r, column=2, value=cust["code"])
    ws.cell(row=r, column=3, value=cust["mst"])
    ws.cell(row=r, column=4, value=level)
    ws.cell(row=r, column=5, value="CHUA CO")
    ws.cell(row=r, column=6, value=None)
    ws.cell(row=r, column=7, value=seq)
    ws.cell(row=r, column=8, value=f'=IF(B{r}="","",(G{r} &"/LT/2025"))')
    ws.cell(row=r, column=9, value=None)
    ws.cell(row=r, column=10, value=None)
    ws.cell(row=r, column=11, value=f'=IF(J{r},"Quý","")')
    ws.cell(row=r, column=12, value=None)
    ws.cell(row=r, column=13, value=f'=IF(B{r}="","",(L{r}+15))')
    ws.cell(row=r, column=14, value=None)


def add_short_term_row(ws, r, cust, via_fallback):
    pairs = []
    if cust["p_level"]:
        pairs.append((SERVICE_NAMES[COL_SERVICE_P], cust["p_level"]))
    if cust["q_level"]:
        pairs.append((SERVICE_NAMES[COL_SERVICE_Q], cust["q_level"]))
    dien_giai = "\n".join(f"{name} - {lvl}" for name, lvl in pairs) or "Chưa được cung cấp"
    no = next_seq(ws, 4, 4, r - 1)
    ws.cell(row=r, column=1, value=f'=IF(D{r}="","",SUBTOTAL(3,$D$4:D{r}))')
    ws.cell(row=r, column=2, value=cust["code"])
    ws.cell(row=r, column=3, value=cust["mst"])
    ws.cell(row=r, column=4, value=no)
    ws.cell(row=r, column=5, value=f'=IF(B{r}="","",(D{r} &"/ST/2025"))')
    ws.cell(row=r, column=6, value=None)
    ws.cell(row=r, column=7, value=dien_giai)
    ws.cell(row=r, column=8, value=None)
    ws.cell(row=r, column=9, value=0.08)
    ws.cell(row=r, column=10, value=f'=H{r}*I{r}')
    ws.cell(row=r, column=11, value=f'=IF(E{r}="","",(E{r}+15))')
    ws.cell(row=r, column=12, value=f'=IF(F{r}="","",(F{r}+15))')
    ws.cell(row=r, column=13, value=None)
    ws.cell(row=r, column=14, value=f'=IF(M{r}="Done",F{r},"")')


def main(client_info_path, template_path):
    customers = read_client_information(client_info_path)

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws_lt = wb[LT_SHEET]
    ws_st = wb[ST_SHEET]

    lt_codes, lt_last = existing_codes(ws_lt, 2)
    st_codes, st_last = existing_codes(ws_st, 2)

    added_lt, added_st = [], []
    skipped_dup_lt, skipped_dup_st = [], []
    skipped_unclassified = []
    both_sheets = []
    fallback_used = []

    for cust in customers:
        code_upper = cust["code"].upper()
        uses_long, uses_short, via_fallback = classify(cust)

        if not uses_long and not uses_short:
            skipped_unclassified.append(cust["code"])
            continue

        if via_fallback:
            fallback_used.append(cust["code"])

        if uses_long:
            if code_upper in lt_codes:
                skipped_dup_lt.append(cust["code"])
            else:
                lt_last += 1
                copy_row_style(ws_lt, lt_last - 1, lt_last, 14)
                add_long_term_row(ws_lt, lt_last, cust, via_fallback)
                lt_codes.add(code_upper)
                added_lt.append(cust["code"])

        if uses_short:
            if code_upper in st_codes:
                skipped_dup_st.append(cust["code"])
            else:
                st_last += 1
                copy_row_style(ws_st, st_last - 1, st_last, 14)
                add_short_term_row(ws_st, st_last, cust, via_fallback)
                st_codes.add(code_upper)
                added_st.append(cust["code"])

        if uses_long and uses_short:
            both_sheets.append(cust["code"])

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"SummaryQuotation_{ts}.xlsx"

    mission_output_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "output"
    )
    os.makedirs(mission_output_dir, exist_ok=True)
    mission_out_path = os.path.join(mission_output_dir, filename)
    wb.save(mission_out_path)

    # Also drop a copy where the chat UI can present it for download.
    delivery_dir = "/mnt/user-data/outputs"
    delivery_path = None
    if os.path.isdir(delivery_dir) or delivery_dir == "/mnt/user-data/outputs":
        try:
            os.makedirs(delivery_dir, exist_ok=True)
            delivery_path = os.path.join(delivery_dir, filename)
            wb.save(delivery_path)
        except Exception:
            delivery_path = None

    print(f"Saved to mission output folder: {mission_out_path}")
    if delivery_path:
        print(f"Copy for download: {delivery_path}")
    print(f"Added to {LT_SHEET}: {added_lt}")
    print(f"Added to {ST_SHEET}: {added_st}")
    print(f"Customers added to BOTH sheets: {both_sheets}")
    print(f"Classified via LOAI fallback (O/P/Q all blank): {fallback_used}")
    print(f"Skipped in {LT_SHEET} (already existed): {skipped_dup_lt}")
    print(f"Skipped in {ST_SHEET} (already existed): {skipped_dup_st}")
    print(f"Skipped entirely (unclassifiable - O/P/Q/LOAI all blank, needs manual review): {skipped_unclassified}")
    return mission_out_path


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 update_summary_quotation.py <client_information.xlsx> <template.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
